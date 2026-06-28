from flask import Flask, render_template, request, send_from_directory
import pandas as pd
from openpyxl import load_workbook
import os
import zipfile

app = Flask(__name__)
LOGS = []
def add_log(message):
    from datetime import datetime
    timestamp = datetime.now().strftime("%H:%M:%S")
    LOGS.append(f"{timestamp}  {message}")
    if len(LOGS) > 20:
        LOGS.pop(0)
        
def render_dashboard(
    rows=0,
    columns=0,
    table=None,
    mapping=None,
    status=None
):

    return render_template(
        "index.html",
        rows=rows,
        columns=columns,
        table=table,
        mapping=mapping or {},
        status=status,
        logs=LOGS
    )                       

def get_oracle_headers():
    

    template_path = os.path.join(
        app.root_path,
        "fbdi_templates",
        "ROG_ScpPlannersImportTemplate.xlsm"
    )

    wb = load_workbook(template_path, read_only=True, data_only=True)

    ws = wb["Planners_"]

    headers = [cell.value for cell in ws[1]][1:]

    headers = [str(h).replace("*", "").strip() for h in headers if h]

    return headers

def auto_map_columns(source_columns):
    
    

    oracle_columns = get_oracle_headers()

    mapping = {}

    synonyms = {

        "srinstancecode": "Source System Code",
        "instancecode": "Source System Code",
        "sourcesystemcode": "Source System Code",
        "sourcecode": "Source System Code",
        "source": "Source System Code",
        "instance": "Source System Code",

        "plannercode": "Planner Code",
        "planner": "Planner Code",
        "plannerid": "Planner Code",
        "plannername": "Planner Code",

        "description": "Description",
        "desc": "Description",
        "details": "Description",

        "employeenumber": "Employee Number",
        "employee": "Employee Number",
        "employeeno": "Employee Number",
        "empno": "Employee Number",
        "employeeid": "Employee Number",
        "empid": "Employee Number",
        "emp": "Employee Number",

        "disabledate": "Disable Date",
        "enddate": "Disable Date",
        "inactivedate": "Disable Date"
    }

    for source in source_columns:

        source_clean = (
            source.lower()
            .replace("_", "")
            .replace(" ", "")
            .replace("-", "")
            .replace(".", "")
        )

        if source_clean in synonyms:
            mapping[source] = synonyms[source_clean]
            continue

        for oracle in oracle_columns:

            oracle_clean = (
                oracle.lower()
                .replace("_", "")
                .replace(" ", "")
                .replace("-", "")
                .replace(".", "")
            )

            if source_clean == oracle_clean:
                mapping[source] = oracle
                break

    return mapping

def normalize_columns(df):

    mapping = auto_map_columns(df.columns.tolist())

    df = df.rename(columns=mapping)

    return df

def generate_fbdi(df):

    oracle_columns = get_oracle_headers()

    fbdi_df = pd.DataFrame(columns=oracle_columns)

    for column in oracle_columns:

        if column in df.columns:
            fbdi_df[column] = df[column]
        else:
            fbdi_df[column] = ""

    return fbdi_df


# ---------------- Home ---------------- #

@app.route("/")
def home():

    return render_dashboard()


# ---------------- Preview ---------------- #

@app.route("/preview", methods=["POST"])
def preview():

    try:
        file = request.files.get("file")
        if file is None or file.filename == "":
         return "<h2>No file selected.</h2>"
        upload_path = os.path.join(
          app.root_path,
          "uploads",
          file.filename
        )
        file.save(upload_path)
        add_log("Excel uploaded")
        df = pd.read_excel(file, dtype=str)

        df = df.fillna("")
        # Remove leading/trailing spaces from column names
        df.columns = df.columns.str.strip()
        df = normalize_columns(df)
        mapping = auto_map_columns(df.columns.tolist())
        add_log("Preview generated")

       
        


        return render_dashboard(
    rows=df.shape[0],
    columns=df.shape[1],
    table=df.head(20).to_html(
        index=False,
        classes="table",
        border=0
    ),
    mapping=mapping
)

    except Exception as e:
        return f"<h2>Error</h2><pre>{e}</pre>"
    # ---------------- Generate FBDI ---------------- #

@app.route("/generate_fbdi", methods=["POST"])
def generate_fbdi_route():

    try:

        upload_folder = os.path.join(app.root_path, "uploads")

        files = os.listdir(upload_folder)

        if len(files) == 0:
            return "<h2>No uploaded file found.</h2>"

        latest_file = max(
            [os.path.join(upload_folder, f) for f in files],
            key=os.path.getctime
        )

        df = pd.read_excel(latest_file, dtype=str)
        df = df.fillna("")

        df.columns = df.columns.str.strip()
        df = normalize_columns(df)

        fbdi_df = generate_fbdi(df)

        output_folder = os.path.join(app.root_path, "output")

        output_file = os.path.join(
            output_folder,
            "ScpPlannersImport.csv"
        )

        fbdi_df.to_csv(
            output_file,
            index=False,
            encoding="utf-8-sig"
        )
        
        add_log("Oracle CSV generated")
        
        zip_file = os.path.join(
            output_folder,
             "ScpPlannersImport.zip"
        )
        
        with zipfile.ZipFile(zip_file, "w", zipfile.ZIP_DEFLATED) as zipf:
           zipf.write(
               output_file,
               arcname="ScpPlannersImport.csv"
        )
        
        add_log("ZIP file generated")

        return render_dashboard(
    status="""
    <span style='color:green;font-weight:bold;'>
    ✅ Oracle FBDI Generated Successfully
    </span>

    <br><br>

    <a href="/download/csv" class="primary-btn">
        ⬇ Download CSV
    </a>

    &nbsp;&nbsp;&nbsp;

    <a href="/download/zip" class="warning-btn">
        ⬇ Download ZIP
    </a>
    """
)

    except Exception as e:
        return f"<h2>Error</h2><pre>{e}</pre>"
    
@app.route("/download/csv")
def download_csv():

    return send_from_directory(
        os.path.join(app.root_path, "output"),
        "ScpPlannersImport.csv",
        as_attachment=True
    )

@app.route("/download/zip")
def download_zip():

    return send_from_directory(
        os.path.join(app.root_path, "output"),
        "ScpPlannersImport.zip",
        as_attachment=True
    )

# ---------------- Validation ---------------- #

@app.route("/validate", methods=["POST"])
def validate():

    try:
        upload_folder = os.path.join(app.root_path, "uploads")

        files = os.listdir(upload_folder)

        if len(files) == 0:
            return "<h2>No uploaded file found.</h2>"

        latest_file = max(
           [os.path.join(upload_folder, f) for f in files],
           key=os.path.getctime
        )

        df = pd.read_excel(latest_file, dtype=str)
        add_log("Validation started")
        df = df.fillna("")

        # Remove spaces from column names
        df.columns = df.columns.str.strip()
        df = normalize_columns(df)
        mapping = auto_map_columns(df.columns.tolist())

        required_columns = [
            "Source System Code",
            "Planner Code",
            "Description",
            "Employee Number",
            "Disable Date"
        ]

        errors = []

        # ---------------------------------------------------
        # 1. Required Columns
        # ---------------------------------------------------

        for column in required_columns:
            if column not in df.columns:
                errors.append(f"❌ Missing Column : {column}")

        if errors:
            return "<br>".join(errors)

        # ---------------------------------------------------
        # 2. Mandatory Fields
        # ---------------------------------------------------
    
        columns_to_check = [
            "Source System Code",
            "Planner Code",
            "Description",
            "Employee Number"
            ]

        for column in columns_to_check:
          blank_rows = df[
            df[column].astype(str).str.strip() == ""
            ]

        for index in blank_rows.index:
          errors.append(
            f"❌ Row {index+2}: {column} cannot be blank"
        )
                
        
   

        # ---------------------------------------------------
        # 3. SR_INSTANCE_CODE should be same
        # ---------------------------------------------------

        unique_source = df["Source System Code"].dropna().astype(str).str.strip().unique()

        if len(unique_source) > 1:
            errors.append("❌ SR_INSTANCE_CODE should be same for all rows.")

        # ---------------------------------------------------
        # 4. PLANNER_CODE should be unique
        # ---------------------------------------------------

        duplicate_planner = df[df["Planner Code"].duplicated(keep=False)]

        for index in duplicate_planner.index:
            errors.append(
                f"❌ Row {index+2}: Duplicate PLANNER_CODE ({df.loc[index, "Planner Code"]})"
            )

        # ---------------------------------------------------
        # 5. EMPLOYEE_NUMBER Numeric (if provided)
        # ---------------------------------------------------


        for index, value in df["Employee Number"].items():

          value = str(value).strip()

          if value == "":
           continue

        if value.endswith(".0"):
          value = value[:-2]

        if len(value) > 30:
         errors.append(
            f"❌ Row {index+2}: EMPLOYEE_NUMBER cannot exceed 30 characters"
        )
    
                    
        # ---------------------------------------------------
        # 6. DISABLE_DATE Validation
        # ---------------------------------------------------

        for index, value in df["Disable Date"].items():

            if pd.notna(value):

                try:
                    pd.to_datetime(value, format="%Y/%m/%d")

                except:
                    errors.append(
                        f"❌ Row {index+2}: Invalid DISABLE_DATE"
                    )

        # ---------------------------------------------------
        # Result
        # ---------------------------------------------------

        if len(errors) == 0:
            add_log("Validation successful")
        

            return render_dashboard(
    status=f"""
    <span style='color:green;font-weight:bold;'>
    ✅ Validation Successful
    </span>

    <br><br>

    Total Records : {len(df)}<br>
    Errors : 0
    """
)

        else:
            add_log("Validation failed")

            return render_dashboard(
    status=f"""
    <span style='color:red;font-weight:bold;'>
    ❌ Validation Failed
    </span>

    <br><br>

    Total Records : {len(df)}<br>
    Total Errors : {len(errors)}

    <hr>

    {'<br>'.join(errors)}
    """
)

    except Exception as e:
        return f"<h2>Error</h2><pre>{e}</pre>"


# ---------------- Run ---------------- #


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)